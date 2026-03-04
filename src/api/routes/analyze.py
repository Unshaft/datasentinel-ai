"""
Route /analyze - Analyse de qualité des données.

Endpoint principal pour analyser un dataset et détecter
les problèmes de qualité.

v0.3 :
- Pipeline async (Profiler séquentiel, Quality checks en parallèle)
- Webhooks déclenchés en background après analyse
- Rate limiting (30/minute par IP)
- Export PDF : GET /analyze/{session_id}/report.pdf
"""

import io
import uuid
from typing import Any

import pandas as pd
from fastapi import APIRouter, BackgroundTasks, HTTPException, Request, status
from fastapi.responses import StreamingResponse

from src.agents.orchestrator import OrchestratorAgent, TaskType
from src.api.limiter import limiter
from src.api.schemas.requests import AnalyzeRequest
from src.api.schemas.responses import (
    AnalyzeResponse,
    ColumnProfileResponse,
    ComparisonResponse,
    DataProfileResponse,
    ErrorResponse,
    IssueResponse,
    SchemaResponse,
    SemanticColumnInfo,
)
from src.core.exceptions import DataSentinelError
from src.core.models import AgentContext
from src.core.dataset_memory import compute_dataset_id, get_dataset_memory_manager
from src.core.stats_manager import get_stats_manager
from src.core.webhook_manager import fire_webhooks
from src.memory.session_store import get_session_store

router = APIRouter(prefix="/analyze", tags=["Analysis"])


def _load_dataframe(body: AnalyzeRequest) -> pd.DataFrame:
    """Charge un DataFrame depuis la requête."""
    if body.data:
        try:
            return pd.DataFrame(body.data)
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Impossible de créer le DataFrame: {str(e)}"
            )

    elif body.file_content:
        try:
            return pd.read_csv(io.StringIO(body.file_content))
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Impossible de parser le CSV: {str(e)}"
            )

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Aucune source de données fournie (data ou file_content requis)"
        )


def _format_profile(profile) -> DataProfileResponse | None:
    """Formate le profil pour la réponse."""
    if profile is None:
        return None

    columns = [
        ColumnProfileResponse(
            name=col.name,
            dtype=col.dtype,
            inferred_type=col.inferred_type,
            count=col.count,
            null_count=col.null_count,
            null_percentage=col.null_percentage,
            unique_count=col.unique_count,
            unique_percentage=col.unique_percentage,
            mean=col.mean,
            std=col.std,
            min=col.min,
            max=col.max,
            sample_values=col.sample_values[:5]
        )
        for col in profile.columns
    ]

    return DataProfileResponse(
        dataset_id=profile.dataset_id,
        row_count=profile.row_count,
        column_count=profile.column_count,
        memory_mb=round(profile.memory_size_bytes / 1024 / 1024, 2),
        total_null_count=profile.total_null_count,
        columns=columns
    )


def _format_issues(issues: list) -> list[IssueResponse]:
    """Formate les issues pour la réponse."""
    return [
        IssueResponse(
            issue_id=issue.issue_id,
            issue_type=issue.issue_type.value,
            severity=issue.severity.value,
            column=issue.column,
            description=issue.description,
            affected_count=issue.affected_count,
            affected_percentage=issue.affected_percentage,
            confidence=issue.confidence,
            details=issue.details
        )
        for issue in issues
    ]


# =============================================================================
# POST /analyze  — Analyse principale
# =============================================================================

@router.post(
    "",
    response_model=AnalyzeResponse,
    responses={
        400: {"model": ErrorResponse, "description": "Données invalides"},
        429: {"description": "Trop de requêtes"},
        500: {"model": ErrorResponse, "description": "Erreur serveur"}
    },
    summary="Analyser un dataset",
    description="""
    Analyse un dataset pour détecter les problèmes de qualité.

    Le système effectue :
    - Profilage statistique complet
    - Détection des valeurs manquantes
    - Détection des anomalies (optionnel)
    - Validation contre les règles métier

    Retourne un score de qualité global et la liste des problèmes détectés.

    **Rate limit** : 30 requêtes / minute par IP.
    """,
)
@limiter.limit("30/minute")
async def analyze_dataset(
    request: Request,
    body: AnalyzeRequest,
    background_tasks: BackgroundTasks,
) -> AnalyzeResponse:
    """Endpoint principal d'analyse de données."""
    try:
        df = _load_dataframe(body)

        if df.empty:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Le dataset est vide"
            )

        orchestrator = OrchestratorAgent()

        dataset_id = compute_dataset_id(df)
        was_known_before = get_dataset_memory_manager().get_entry(dataset_id) is not None
        context = AgentContext(
            session_id=f"session_{uuid.uuid4().hex[:12]}",
            dataset_id=dataset_id
        )

        options = {
            "detect_anomalies": body.detect_anomalies,
            "detect_drift": body.detect_drift,
        }

        if body.custom_rules:
            from src.memory.chroma_store import get_chroma_store
            store = get_chroma_store()
            for i, rule_text in enumerate(body.custom_rules):
                store.add_rule(
                    rule_id=f"custom_{context.session_id}_{i}",
                    rule_text=rule_text,
                    rule_type="custom"
                )

        # Pipeline : adaptatif (ReAct) si include_reasoning, sinon async standard
        if body.include_reasoning:
            context = await orchestrator.run_pipeline_adaptive(
                context, df,
                task_type=TaskType.ANALYZE,
                **options
            )
        else:
            context = await orchestrator.run_pipeline_async(
                context, df,
                task_type=TaskType.ANALYZE,
                **options
            )

        issues_by_severity: dict[str, int] = {}
        for issue in context.issues:
            sev = issue.severity.value
            issues_by_severity[sev] = issues_by_severity.get(sev, 0) + 1

        escalation_reasons = []
        if context.metadata.get("needs_human_review"):
            critical_count = issues_by_severity.get("critical", 0)
            if critical_count > 0:
                escalation_reasons.append(f"{critical_count} problème(s) critique(s)")
            low_confidence = [i for i in context.issues if i.confidence < 0.5]
            if low_confidence:
                escalation_reasons.append(
                    f"{len(low_confidence)} détection(s) à faible confiance"
                )

        response = AnalyzeResponse(
            session_id=context.session_id,
            dataset_id=context.dataset_id,
            status=context.metadata.get("final_status", "completed"),
            quality_score=context.metadata.get("quality_score", 100),
            processing_time_ms=context.metadata.get("processing_time_ms", 0),
            summary=context.metadata.get("summary", ""),
            profile=_format_profile(context.profile),
            issues=_format_issues(context.issues),
            issues_by_severity=issues_by_severity,
            column_scores=context.metadata.get("column_scores", {}),
            semantic_types=context.metadata.get("semantic_types") or None,
            domain_agent=context.metadata.get("domain_name"),
            reflect_flags=context.metadata.get("reflect_flags", []),
            reasoning_steps=context.metadata.get("reasoning_steps", []),
            needs_human_review=context.metadata.get("needs_human_review", False),
            escalation_reasons=escalation_reasons,
            dataset_memory=None,  # rempli après record_session ci-dessous
        )

        # Persistance best-effort (contexte + DataFrame original)
        try:
            store = get_session_store()
            store.save(context.session_id, context)
            store.save_dataframe(context.session_id, df)
        except Exception:
            pass

        # Stats (best-effort)
        try:
            issue_types = [iss.issue_type.value for iss in context.issues]
            get_stats_manager().record_session(
                context.metadata.get("quality_score", 100),
                issue_types,
            )
        except Exception:
            pass

        # Dataset memory (best-effort)
        try:
            mem = get_dataset_memory_manager()
            mem.record_session(
                dataset_id=context.dataset_id,
                session_id=context.session_id,
                quality_score=context.metadata.get("quality_score", 100),
                issues=context.issues,
            )
            mem_info = mem.get_memory_info(context.dataset_id, was_known_before)
            from src.api.schemas.responses import DatasetMemoryInfo
            response.dataset_memory = DatasetMemoryInfo(**mem_info)
        except Exception:
            pass

        # Webhooks en background (non-bloquant)
        background_tasks.add_task(
            fire_webhooks,
            "analysis.completed",
            {
                "session_id": context.session_id,
                "dataset_id": context.dataset_id,
                "quality_score": context.metadata.get("quality_score", 100),
                "status": context.metadata.get("final_status", "completed"),
                "issues_count": len(context.issues),
                "needs_human_review": context.metadata.get("needs_human_review", False),
            },
        )

        return response

    except HTTPException:
        raise
    except DataSentinelError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=e.message
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de l'analyse: {str(e)}"
        )


# =============================================================================
# GET /analyze/{session_id}  — Récupération session
# =============================================================================

@router.get(
    "/{session_id}",
    response_model=AnalyzeResponse,
    responses={
        404: {"model": ErrorResponse, "description": "Session non trouvée"}
    },
    summary="Récupérer les résultats d'une analyse",
    description="Récupère les résultats d'une analyse précédente par son ID de session.",
)
async def get_analysis_results(session_id: str) -> AnalyzeResponse:
    """Récupère les résultats d'une session existante."""
    context = get_session_store().load(session_id)
    if context is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Session '{session_id}' introuvable ou expirée."
        )

    issues_by_severity: dict[str, int] = {}
    for issue in context.issues:
        sev = issue.severity.value
        issues_by_severity[sev] = issues_by_severity.get(sev, 0) + 1

    return AnalyzeResponse(
        session_id=context.session_id,
        dataset_id=context.dataset_id,
        status=context.metadata.get("final_status", "completed"),
        quality_score=context.metadata.get("quality_score", 100),
        processing_time_ms=context.metadata.get("processing_time_ms", 0),
        summary=context.metadata.get("summary", ""),
        profile=_format_profile(context.profile),
        issues=_format_issues(context.issues),
        issues_by_severity=issues_by_severity,
        column_scores=context.metadata.get("column_scores", {}),
        semantic_types=context.metadata.get("semantic_types") or None,
        domain_agent=context.metadata.get("domain_name"),
        reflect_flags=context.metadata.get("reflect_flags", []),
        reasoning_steps=context.metadata.get("reasoning_steps", []),
        needs_human_review=context.metadata.get("needs_human_review", False),
        escalation_reasons=[]
    )


# =============================================================================
# GET /analyze/{session_id}/report.pdf  — Export PDF
# =============================================================================

@router.get(
    "/{session_id}/report.pdf",
    summary="Exporter le rapport PDF",
    description="""
    Génère et télécharge un rapport PDF complet pour une session existante.

    Le rapport contient :
    - Informations de session (ID, score, statut, date)
    - Score de qualité (coloré selon le niveau)
    - Tableau des problèmes détectés
    - Résumé du profil des données
    """,
    response_class=StreamingResponse,
    responses={
        200: {"content": {"application/pdf": {}}, "description": "Fichier PDF"},
        404: {"model": ErrorResponse, "description": "Session non trouvée"},
        503: {"description": "reportlab non installé"},
    },
)
async def export_pdf_report(session_id: str) -> StreamingResponse:
    """Génère un rapport PDF pour une session existante."""
    context = get_session_store().load(session_id)
    if context is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Session '{session_id}' introuvable ou expirée."
        )

    try:
        pdf_bytes = _generate_pdf(context)
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="reportlab n'est pas installé. Installez-le avec : pip install reportlab",
        )

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=rapport_{session_id}.pdf"
        },
    )


# =============================================================================
# GET /analyze/{session_id}/corrections  — Plan de corrections (v0.4)
# =============================================================================

@router.get(
    "/{session_id}/corrections",
    summary="Plan de corrections automatiques",
    description="""
    Retourne un plan de corrections basé sur les issues détectées.

    Pour chaque problème, indique l'action recommandée, si elle est
    applicable automatiquement, et l'impact estimé.
    """,
    responses={
        404: {"model": ErrorResponse, "description": "Session non trouvée"},
    },
)
async def get_corrections_plan(session_id: str) -> dict:
    """Génère un plan de corrections depuis les issues d'une session."""
    context = get_session_store().load(session_id)
    if context is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Session '{session_id}' introuvable ou expirée.",
        )

    auto: list[dict] = []
    manual: list[dict] = []

    from src.core.models import IssueType

    _AUTO_ACTIONS = {
        IssueType.MISSING_VALUES:       ("impute_median / impute_mode", True),
        IssueType.DUPLICATE:            ("delete_duplicate_rows", True),
        IssueType.TYPE_MISMATCH:        ("cast_to_dominant_type", True),
        IssueType.FORMAT_ERROR:         ("standardize_format", True),
        IssueType.ANOMALY:              ("flag_for_review", False),
        IssueType.CONSTRAINT_VIOLATION: ("manual_review", False),
        IssueType.DRIFT:                ("manual_review", False),
        IssueType.OUTLIER:              ("clip_values", False),
        IssueType.INCONSISTENCY:        ("manual_review", False),
    }

    for issue in context.issues:
        action, is_auto = _AUTO_ACTIONS.get(issue.issue_type, ("manual_review", False))
        entry = {
            "issue_id": issue.issue_id,
            "issue_type": issue.issue_type.value,
            "severity": issue.severity.value,
            "column": issue.column,
            "description": issue.description,
            "recommended_action": action,
            "auto_applicable": is_auto,
            "affected_count": issue.affected_count,
            "affected_percentage": issue.affected_percentage,
        }
        (auto if is_auto else manual).append(entry)

    return {
        "session_id": session_id,
        "quality_score": context.metadata.get("quality_score", 100),
        "total_issues": len(context.issues),
        "auto_corrections": auto,
        "manual_reviews": manual,
        "estimated_score_after_auto": min(
            100.0,
            context.metadata.get("quality_score", 100)
            + sum(
                {"critical": 15, "high": 10, "medium": 5, "low": 2}.get(c["severity"], 0)
                for c in auto
            ),
        ),
    }


# =============================================================================
# GET /analyze/{session_id}/report.xlsx  — Export Excel (v0.4)
# =============================================================================

@router.get(
    "/{session_id}/report.xlsx",
    summary="Exporter le rapport Excel",
    description="""
    Génère un rapport Excel multi-onglets :
    - **Résumé** : score, statut, metadata
    - **Issues** : tableau de tous les problèmes
    - **Profil** : statistiques par colonne
    - **Score par colonne** : score qualité individuel
    """,
    response_class=StreamingResponse,
    responses={
        200: {
            "content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}},
            "description": "Fichier Excel",
        },
        404: {"model": ErrorResponse, "description": "Session non trouvée"},
        503: {"description": "openpyxl non installé"},
    },
)
async def export_excel_report(session_id: str) -> StreamingResponse:
    """Génère un rapport Excel pour une session existante."""
    context = get_session_store().load(session_id)
    if context is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Session '{session_id}' introuvable ou expirée.",
        )

    try:
        xlsx_bytes = _generate_excel(context)
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="openpyxl n'est pas installé. Installez-le avec : pip install openpyxl",
        )

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=rapport_{session_id}.xlsx"
        },
    )


# =============================================================================
# POST /analyze/{session_id}/apply-corrections  — Application des corrections (v0.5)
# =============================================================================

# Pseudo-nulls à remplacer par NaN avant imputation
_PSEUDO_NULL_TOKENS: frozenset = frozenset({
    "n/a", "na", "null", "none", "-", "--", "?", "??",
    "nan", "#n/a", "#na", "missing", "unknown", "inconnu",
    "nd", "nr", "nc", "n.a.", "n.a", "not available",
})


@router.post(
    "/{session_id}/apply-corrections",
    summary="Appliquer les corrections automatiques",
    description="""
    Applique les corrections automatiques au dataset original et retourne un CSV propre.

    Corrections appliquées :
    - **Doublons** : suppression (keep=first)
    - **Valeurs manquantes** : remplacement pseudo-nulls → NaN, puis imputation
      (médiane pour les numériques, mode pour les catégorielles)
    - **Erreurs de format** : suppression des espaces parasites
    - **Type mismatch** : tentative de cast numérique

    Les données originales doivent être disponibles en session.
    Si elles ont expiré, relancez l'analyse.
    """,
    response_class=StreamingResponse,
    responses={
        200: {"content": {"text/csv": {}}, "description": "CSV corrigé"},
        404: {"model": ErrorResponse, "description": "Session non trouvée"},
        422: {"model": ErrorResponse, "description": "Données originales expirées"},
    },
)
async def apply_corrections(session_id: str) -> StreamingResponse:
    """Applique les corrections auto et retourne le CSV corrigé."""
    from src.core.models import IssueType

    store = get_session_store()
    context = store.load(session_id)
    if context is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Session '{session_id}' introuvable ou expirée.",
        )

    df = store.load_dataframe(session_id)
    if df is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                "Données originales non disponibles "
                "(session trop ancienne ou analyse via JSON). "
                "Relancez l'analyse via POST /upload."
            ),
        )

    df = df.copy()
    rows_before = len(df)
    corrections: list[str] = []

    # ── 1. Suppression des doublons ───────────────────────────────────────────
    if any(iss.issue_type == IssueType.DUPLICATE for iss in context.issues):
        before = len(df)
        df = df.drop_duplicates(keep="first")
        removed = before - len(df)
        if removed:
            corrections.append(f"Doublons: {removed} ligne(s) supprimée(s)")

    # ── 2. Corrections par colonne (chaque colonne traitée une seule fois) ────
    processed: set[str] = set()
    for issue in context.issues:
        col = issue.column
        if col is None or col not in df.columns or col in processed:
            continue

        if issue.issue_type == IssueType.MISSING_VALUES:
            # Remplace les pseudo-nulls par NaN
            mask = df[col].astype(str).str.lower().str.strip().isin(_PSEUDO_NULL_TOKENS)
            if mask.any():
                df.loc[mask, col] = pd.NA

            # Imputation
            if pd.api.types.is_numeric_dtype(df[col]):
                median_val = df[col].median()
                if pd.notna(median_val):
                    df[col] = df[col].fillna(median_val)
                    corrections.append(f"{col}: nulls → médiane ({median_val:.2f})")
            else:
                mode_vals = df[col].dropna().mode()
                if not mode_vals.empty:
                    df[col] = df[col].fillna(mode_vals.iloc[0])
                    corrections.append(f"{col}: nulls → mode ({mode_vals.iloc[0]})")
            processed.add(col)

        elif issue.issue_type == IssueType.FORMAT_ERROR:
            df[col] = df[col].astype(str).str.strip()
            corrections.append(f"{col}: espaces supprimés")
            processed.add(col)

        elif issue.issue_type == IssueType.TYPE_MISMATCH:
            numeric_attempt = pd.to_numeric(df[col], errors="coerce")
            fill_rate = numeric_attempt.notna().sum() / max(len(df), 1)
            if fill_rate >= 0.9:
                df[col] = numeric_attempt
                corrections.append(f"{col}: type → numérique")
            processed.add(col)

    # ── 3. Sérialisation CSV ──────────────────────────────────────────────────
    buf = io.BytesIO()
    df.to_csv(buf, index=False)
    csv_bytes = buf.getvalue()

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=corrected_{session_id}.csv",
            "X-Rows-Before": str(rows_before),
            "X-Rows-After": str(len(df)),
            "X-Corrections-Count": str(len(corrections)),
        },
    )


# =============================================================================
# GET /analyze/{session_id}/comparison  — Comparison avant/après (v0.6 — F19)
# =============================================================================

@router.get(
    "/{session_id}/comparison",
    response_model=ComparisonResponse,
    summary="Comparer la qualité avant/après corrections",
    description="""
    Applique les corrections automatiques en mémoire et mesure l'impact
    sur le score de qualité, sans modifier les données persistées.

    Nécessite que le DataFrame original soit disponible en session
    (créé via POST /upload ou POST /analyze avec des données).
    """,
    responses={
        404: {"model": ErrorResponse, "description": "Session non trouvée"},
        422: {"model": ErrorResponse, "description": "Données originales expirées"},
    },
)
async def get_comparison(session_id: str) -> ComparisonResponse:
    """Compare les scores de qualité avant et après corrections automatiques."""
    from src.agents.quality import QualityAgent
    from src.core.models import AgentContext as Ctx, IssueType

    store = get_session_store()
    context = store.load(session_id)
    if context is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Session '{session_id}' introuvable ou expirée.",
        )

    df_orig = store.load_dataframe(session_id)
    if df_orig is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                "Données originales non disponibles. "
                "Relancez l'analyse via POST /upload."
            ),
        )

    score_before: float = context.metadata.get("quality_score", 100)
    issues_before = {iss.issue_type.value for iss in context.issues}

    # Applique les mêmes corrections qu'apply-corrections
    df_corrected = df_orig.copy()

    if any(iss.issue_type == IssueType.DUPLICATE for iss in context.issues):
        df_corrected = df_corrected.drop_duplicates(keep="first")

    processed: set[str] = set()
    for issue in context.issues:
        col = issue.column
        if col is None or col not in df_corrected.columns or col in processed:
            continue
        if issue.issue_type == IssueType.MISSING_VALUES:
            mask = df_corrected[col].astype(str).str.lower().str.strip().isin(_PSEUDO_NULL_TOKENS)
            if mask.any():
                df_corrected.loc[mask, col] = pd.NA
            if pd.api.types.is_numeric_dtype(df_corrected[col]):
                m = df_corrected[col].median()
                if pd.notna(m):
                    df_corrected[col] = df_corrected[col].fillna(m)
            else:
                mv = df_corrected[col].dropna().mode()
                if not mv.empty:
                    df_corrected[col] = df_corrected[col].fillna(mv.iloc[0])
            processed.add(col)
        elif issue.issue_type == IssueType.FORMAT_ERROR:
            df_corrected[col] = df_corrected[col].astype(str).str.strip()
            processed.add(col)
        elif issue.issue_type == IssueType.TYPE_MISMATCH:
            num = pd.to_numeric(df_corrected[col], errors="coerce")
            if num.notna().sum() / max(len(df_corrected), 1) >= 0.9:
                df_corrected[col] = num
            processed.add(col)

    # Re-run quality uniquement sur le df corrigé
    try:
        quality_agent = QualityAgent()
        ctx_copy = Ctx(
            session_id=f"cmp_{session_id}",
            dataset_id=context.dataset_id,
        )
        ctx_copy.profile = context.profile
        ctx_copy = quality_agent.execute(
            ctx_copy, df_corrected, detect_anomalies=False, detect_drift=False
        )

        # Calcul du score après
        from src.agents.orchestrator import OrchestratorAgent
        orch = OrchestratorAgent()
        score_after = orch._calculate_quality_score(ctx_copy)
    except Exception:
        # Fallback : estimation basique
        auto_types = {IssueType.DUPLICATE, IssueType.MISSING_VALUES,
                      IssueType.FORMAT_ERROR, IssueType.TYPE_MISMATCH}
        auto_fixed = [i for i in context.issues if i.issue_type in auto_types]
        bonus = sum(
            {"critical": 15, "high": 10, "medium": 5, "low": 2}.get(i.severity.value, 0)
            for i in auto_fixed
        )
        score_after = min(100.0, score_before + bonus)
        ctx_copy = Ctx(session_id=f"cmp_{session_id}", dataset_id=context.dataset_id)

    issues_after = {iss.issue_type.value for iss in ctx_copy.issues}
    issues_removed = sorted(issues_before - issues_after)
    issues_remaining = sorted(issues_before & issues_after)

    # Colonnes dont le score a augmenté
    col_scores_before: dict[str, float] = context.metadata.get("column_scores", {})
    col_scores_after: dict[str, float] = ctx_copy.metadata.get("column_scores", {})
    columns_improved = [
        col for col in col_scores_before
        if col_scores_after.get(col, col_scores_before[col]) > col_scores_before[col]
    ]

    return ComparisonResponse(
        session_id=session_id,
        score_before=round(score_before, 2),
        score_after=round(score_after, 2),
        delta=round(score_after - score_before, 2),
        issues_removed=issues_removed,
        issues_remaining=issues_remaining,
        columns_improved=columns_improved,
    )


# =============================================================================
# GET /analyze/{session_id}/schema  — Export schéma sémantique (v0.8 — F29)
# =============================================================================

@router.get(
    "/{session_id}/schema",
    response_model=SchemaResponse,
    summary="Exporter le schéma sémantique détecté",
    description="""
    Retourne le schéma sémantique du dataset analysé.

    Fusionne le profil technique (types pandas) et les types sémantiques LLM
    (disponibles si ENABLE_LLM_CHECKS=true au moment de l'analyse).

    - `inferred_type` : type technique détecté par le ProfilingAgent
    - `semantic_type` : nature métier détectée par le SemanticProfilerAgent (LLM)
    - `confidence` : confiance du LLM sur la classification sémantique
    """,
    responses={
        404: {"model": ErrorResponse, "description": "Session non trouvée"},
    },
)
async def get_schema(session_id: str) -> SchemaResponse:
    """Retourne le schéma sémantique d'une session analysée."""
    context = get_session_store().load(session_id)
    if context is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Session '{session_id}' introuvable ou expirée.",
        )

    semantic_types: dict = context.metadata.get("semantic_types", {})
    col_infos: list[SemanticColumnInfo] = []

    if context.profile:
        for col in context.profile.columns:
            sem = semantic_types.get(col.name, {})
            col_infos.append(SemanticColumnInfo(
                name=col.name,
                dtype=col.dtype,
                inferred_type=col.inferred_type,
                semantic_type=sem.get("semantic_type"),
                confidence=sem.get("confidence"),
                language=sem.get("language"),
                pattern=sem.get("pattern"),
                notes=sem.get("notes"),
                null_percentage=col.null_percentage,
                unique_count=col.unique_count,
                sample_values=col.sample_values[:5],
            ))
    else:
        # Pas de profil → colonnes vides avec semantic_types si disponibles
        for col_name, sem in semantic_types.items():
            col_infos.append(SemanticColumnInfo(
                name=col_name,
                dtype="unknown",
                inferred_type="unknown",
                semantic_type=sem.get("semantic_type"),
                confidence=sem.get("confidence"),
                language=sem.get("language"),
                pattern=sem.get("pattern"),
                notes=sem.get("notes"),
            ))

    total = len(col_infos)
    covered = sum(1 for c in col_infos if c.semantic_type is not None)
    semantic_coverage = round(covered / total * 100, 1) if total > 0 else 0.0

    return SchemaResponse(
        session_id=session_id,
        dataset_id=context.dataset_id,
        columns=col_infos,
        semantic_coverage=semantic_coverage,
    )


def _generate_excel(context: AgentContext) -> bytes:
    """Génère un rapport Excel multi-onglets depuis un AgentContext."""
    from datetime import datetime

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_align = Alignment(horizontal="center", vertical="center")

    sev_fills = {
        "critical": PatternFill("solid", fgColor="E74C3C"),
        "high":     PatternFill("solid", fgColor="E67E22"),
        "medium":   PatternFill("solid", fgColor="F39C12"),
        "low":      PatternFill("solid", fgColor="3498DB"),
    }
    white_font = Font(color="FFFFFF", bold=True)

    def _header_row(ws, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align

    def _autofit(ws) -> None:
        for col_cells in ws.columns:
            length = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 60)

    quality_score = context.metadata.get("quality_score", 100)

    # ── Sheet 1 : Résumé ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    rows_summary = [
        ["DataSentinel AI — Rapport de qualité"],
        [],
        ["Session ID", context.session_id],
        ["Dataset ID", context.dataset_id],
        ["Statut", context.metadata.get("final_status", "completed").upper()],
        ["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Temps traitement", f"{context.metadata.get('processing_time_ms', 0)} ms"],
        [],
        ["Score de qualité", f"{quality_score:.1f} / 100"],
        ["Résumé", context.metadata.get("summary", "")],
        [],
        ["Problèmes total", len(context.issues)],
    ]
    for row in rows_summary:
        ws_summary.append(row)
    ws_summary["A1"].font = Font(bold=True, size=14)
    _autofit(ws_summary)

    # ── Sheet 2 : Issues ──────────────────────────────────────────────────────
    ws_issues = wb.create_sheet("Issues")
    _header_row(ws_issues, ["Type", "Sévérité", "Colonne", "Description", "Affecté #", "Affecté %", "Confiance"])
    for issue in context.issues:
        ws_issues.append([
            issue.issue_type.value.replace("_", " ").title(),
            issue.severity.value.upper(),
            issue.column or "—",
            issue.description,
            issue.affected_count,
            f"{issue.affected_percentage:.1f}%",
            f"{issue.confidence:.0%}",
        ])
        sev = issue.severity.value
        if sev in sev_fills:
            row = ws_issues.max_row
            ws_issues.cell(row, 2).fill = sev_fills[sev]
            ws_issues.cell(row, 2).font = white_font
    _autofit(ws_issues)

    # ── Sheet 3 : Profil colonnes ─────────────────────────────────────────────
    ws_profile = wb.create_sheet("Profil")
    if context.profile:
        _header_row(ws_profile, [
            "Colonne", "Type", "Type inféré", "Count", "Nulls", "Nulls %",
            "Uniques", "Uniques %", "Min", "Max", "Moyenne", "Écart-type",
        ])
        for col in context.profile.columns:
            ws_profile.append([
                col.name, col.dtype, col.inferred_type,
                col.count, col.null_count,
                f"{col.null_percentage:.1f}%",
                col.unique_count,
                f"{col.unique_percentage:.1f}%",
                col.min, col.max,
                round(col.mean, 4) if col.mean is not None else "",
                round(col.std, 4) if col.std is not None else "",
            ])
    _autofit(ws_profile)

    # ── Sheet 4 : Score par colonne ───────────────────────────────────────────
    ws_scores = wb.create_sheet("Score par colonne")
    _header_row(ws_scores, ["Colonne", "Score", "Niveau"])
    col_scores = context.metadata.get("column_scores", {})
    for col_name, score in sorted(col_scores.items(), key=lambda x: x[1]):
        level = "Excellent" if score >= 90 else "Bon" if score >= 70 else "Moyen" if score >= 50 else "Mauvais"
        ws_scores.append([col_name, score, level])
        row = ws_scores.max_row
        fill_color = (
            "27AE60" if score >= 90
            else "F39C12" if score >= 70
            else "E67E22" if score >= 50
            else "E74C3C"
        )
        ws_scores.cell(row, 2).fill = PatternFill("solid", fgColor=fill_color)
        ws_scores.cell(row, 2).font = Font(color="FFFFFF", bold=True)
    _autofit(ws_scores)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_pdf(context: AgentContext) -> bytes:
    """Génère un rapport PDF professionnel depuis un AgentContext."""
    from datetime import datetime

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    elements: list = []

    quality_score = context.metadata.get("quality_score", 100)
    if quality_score >= 80:
        score_color = colors.HexColor("#27ae60")
    elif quality_score >= 60:
        score_color = colors.HexColor("#e67e22")
    else:
        score_color = colors.HexColor("#e74c3c")

    # --- Titre ---
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=22,
        textColor=colors.HexColor("#2c3e50"),
    )
    elements.append(Paragraph("DataSentinel AI", title_style))
    elements.append(Paragraph("Rapport de Qualité des Données", styles["Heading2"]))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    elements.append(Spacer(1, 0.3 * cm))

    # --- Infos session ---
    info_data = [
        ["Session ID", context.session_id],
        ["Dataset ID", context.dataset_id],
        ["Statut", context.metadata.get("final_status", "completed").upper()],
        ["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Temps traitement", f"{context.metadata.get('processing_time_ms', 0)} ms"],
    ]
    info_table = Table(info_data, colWidths=[5 * cm, 11 * cm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#2c3e50")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.5 * cm))

    # --- Score de qualité ---
    elements.append(Paragraph("Score de Qualité", styles["Heading2"]))
    score_style = ParagraphStyle(
        "Score",
        parent=styles["Normal"],
        fontSize=32,
        textColor=score_color,
        leading=38,
    )
    elements.append(Paragraph(f"{quality_score:.1f} / 100", score_style))
    if summary := context.metadata.get("summary"):
        elements.append(Spacer(1, 0.2 * cm))
        elements.append(Paragraph(summary, styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    # --- Problèmes détectés ---
    if context.issues:
        elements.append(
            Paragraph(f"Problèmes Détectés ({len(context.issues)})", styles["Heading2"])
        )
        elements.append(Spacer(1, 0.2 * cm))

        sev_colors = {
            "critical": colors.HexColor("#e74c3c"),
            "high": colors.HexColor("#e67e22"),
            "medium": colors.HexColor("#f39c12"),
            "low": colors.HexColor("#3498db"),
        }

        issue_data = [["Type", "Sévérité", "Colonne", "Description", "Affecté %"]]
        for issue in context.issues[:50]:
            desc = issue.description
            if len(desc) > 55:
                desc = desc[:55] + "…"
            issue_data.append([
                issue.issue_type.value.replace("_", " ").title(),
                issue.severity.value.upper(),
                issue.column or "—",
                desc,
                f"{issue.affected_percentage:.1f}%",
            ])

        issue_table = Table(
            issue_data,
            colWidths=[3.2 * cm, 2.3 * cm, 2.8 * cm, 6.5 * cm, 2.2 * cm],
        )

        table_style_cmds: list = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ]
        for i, issue in enumerate(context.issues[:50], start=1):
            sev = issue.severity.value
            if sev in sev_colors:
                table_style_cmds.append(("TEXTCOLOR", (1, i), (1, i), sev_colors[sev]))
                table_style_cmds.append(("FONTNAME", (1, i), (1, i), "Helvetica-Bold"))

        issue_table.setStyle(TableStyle(table_style_cmds))
        elements.append(issue_table)
    else:
        elements.append(
            Paragraph("✓ Aucun problème de qualité détecté", styles["Normal"])
        )

    # --- Profil ---
    if context.profile:
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(Paragraph("Profil des Données", styles["Heading2"]))
        profile_data = [
            ["Lignes", str(context.profile.row_count)],
            ["Colonnes", str(context.profile.column_count)],
            ["Valeurs manquantes", str(context.profile.total_null_count)],
            ["Mémoire", f"{round(context.profile.memory_size_bytes / 1024 / 1024, 2)} MB"],
        ]
        profile_table = Table(profile_data, colWidths=[5 * cm, 11 * cm])
        profile_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(profile_table)

    doc.build(elements)
    return buf.getvalue()
